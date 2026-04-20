from __future__ import annotations

import json
import os
import tempfile
from dataclasses import asdict, dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from monitor_app.config import normalize_monitor_id, parse_monitor_accounts_payload
from monitor_app.secrets.provider import EncryptedFileSecretProvider
from monitor_app.secrets.refs import (
    access_control_secret_ref,
    child_account_secret_ref,
    main_account_secret_ref,
    telegram_secret_ref,
)


TEMPLATE_VERSION = "2"
REQUIRED_COLUMNS = ("main_id", "main_name", "account_id", "name", "api_key", "api_secret")
OPTIONAL_COLUMNS = ("uid", "use_testnet", "rest_base_url", "ws_base_url")
SUPPORTED_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS
SETTINGS_REQUIRED_COLUMNS = ("key", "value")
SETTINGS_OPTIONAL_COLUMNS = ("notes",)
SUPPORTED_SETTINGS_COLUMNS = SETTINGS_REQUIRED_COLUMNS + SETTINGS_OPTIONAL_COLUMNS
TRUE_VALUES = {"1", "true", "yes", "on"}
FALSE_VALUES = {"0", "false", "no", "off"}


class AccountImportError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class SettingSpec:
    key: str
    secret_ref: str
    target: str
    ref_field: str
    plaintext_field: str
    env_var: str = ""


DEFAULT_SETTING_SPECS: dict[str, SettingSpec] = {
    "telegram.bot_token": SettingSpec(
        key="telegram.bot_token",
        secret_ref=telegram_secret_ref("bot_token"),
        target="env",
        ref_field="TG_BOT_TOKEN_SECRET_REF",
        plaintext_field="TG_BOT_TOKEN",
        env_var="TG_BOT_TOKEN_SECRET_REF",
    ),
    "telegram.chat_id": SettingSpec(
        key="telegram.chat_id",
        secret_ref=telegram_secret_ref("chat_id"),
        target="env",
        ref_field="TG_CHAT_ID_SECRET_REF",
        plaintext_field="TG_CHAT_ID",
        env_var="TG_CHAT_ID_SECRET_REF",
    ),
    "access_control.guest_password": SettingSpec(
        key="access_control.guest_password",
        secret_ref=access_control_secret_ref("guest_password"),
        target="access_control",
        ref_field="guest_password_secret_ref",
        plaintext_field="guest_password",
    ),
    "access_control.admin_password": SettingSpec(
        key="access_control.admin_password",
        secret_ref=access_control_secret_ref("admin_password"),
        target="access_control",
        ref_field="admin_password_secret_ref",
        plaintext_field="admin_password",
    ),
    "access_control.session_secret": SettingSpec(
        key="access_control.session_secret",
        secret_ref=access_control_secret_ref("session_secret"),
        target="access_control",
        ref_field="session_secret_secret_ref",
        plaintext_field="session_secret",
    ),
}


@dataclass(frozen=True, slots=True)
class ImportResult:
    file_name: str
    main_account_count: int
    account_count: int
    updated_settings_keys: tuple[str, ...] = ()
    mode: str = "replace_all"

    def as_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True, slots=True)
class ParsedExcelImport:
    payload: dict[str, Any]
    import_result: ImportResult
    settings_updates: dict[str, str]
    template_version: str


def parse_accounts_excel(
    content: bytes,
    *,
    filename: str,
    setting_specs: dict[str, SettingSpec] | None = None,
) -> ParsedExcelImport:
    if not content:
        raise AccountImportError("Excel file is empty")

    try:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - library-specific failures vary
        raise AccountImportError("Failed to read Excel workbook") from exc

    try:
        if not workbook.worksheets:
            raise AccountImportError("Excel workbook must contain at least one worksheet")

        specs = setting_specs or DEFAULT_SETTING_SPECS
        template_version = _parse_template_version(workbook)
        if template_version not in {"legacy", TEMPLATE_VERSION}:
            raise AccountImportError(f"Unsupported template_version: {template_version}")

        accounts_sheet = workbook["accounts"] if "accounts" in workbook.sheetnames else workbook.worksheets[0]
        payload, result = _parse_accounts_worksheet(accounts_sheet, filename=filename)
        settings_updates = (
            _parse_settings_worksheet(workbook["settings"], setting_specs=specs)
            if "settings" in workbook.sheetnames
            else {}
        )
        if "settings" in workbook.sheetnames and template_version == "legacy":
            raise AccountImportError("settings worksheet requires a supported template_version")
        if not payload["main_accounts"] and not settings_updates:
            raise AccountImportError("Excel workbook must include at least one account row or one settings update")
        import_mode = "settings_only" if not payload["main_accounts"] and settings_updates else result.mode

        return ParsedExcelImport(
            payload=payload,
            import_result=ImportResult(
                file_name=result.file_name,
                main_account_count=result.main_account_count,
                account_count=result.account_count,
                updated_settings_keys=tuple(sorted(settings_updates.keys())),
                mode=import_mode,
            ),
            settings_updates=settings_updates,
            template_version=template_version,
        )
    finally:
        workbook.close()


def write_monitor_accounts_payload(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    _write_text_atomic(path, json.dumps(payload, ensure_ascii=False, indent=2) + "\n")


def write_text_atomic(path: Path, content: str) -> None:
    _write_text_atomic(path, content)


def materialize_monitor_accounts_secret_refs(
    payload: dict[str, Any],
    *,
    secret_provider: EncryptedFileSecretProvider | None = None,
    secret_store: dict[str, str] | None = None,
) -> tuple[dict[str, Any], set[str], dict[str, str]]:
    main_accounts = payload.get("main_accounts") if isinstance(payload, dict) else None
    if not isinstance(main_accounts, list):
        raise AccountImportError("Monitor accounts payload must contain a main_accounts array")

    if secret_store is not None:
        next_store = {str(key): str(value) for key, value in dict(secret_store).items()}
    elif secret_provider is not None:
        next_store = secret_provider.dump_store()
    else:
        raise ValueError("secret_provider or secret_store is required")

    rendered_main_accounts: list[dict[str, Any]] = []
    used_secret_refs: set[str] = set()
    for raw_main_account in main_accounts:
        if not isinstance(raw_main_account, dict):
            raise AccountImportError("Each main account entry must be an object")
        main_id = str(raw_main_account.get("main_id") or "").strip().lower()
        if not main_id:
            raise AccountImportError("main_id is required")

        transfer_api_key = str(raw_main_account.get("transfer_api_key") or "").strip()
        transfer_api_secret = str(raw_main_account.get("transfer_api_secret") or "").strip()
        rendered_main_account: dict[str, Any] = {
            "main_id": main_id,
            "name": str(raw_main_account.get("name") or "").strip(),
            "transfer_uid": str(raw_main_account.get("transfer_uid") or "").strip(),
            "children": [],
        }
        if transfer_api_key:
            transfer_api_key_ref = main_account_secret_ref(main_id, "transfer_api_key")
            next_store[transfer_api_key_ref] = transfer_api_key
            rendered_main_account["transfer_api_key_secret_ref"] = transfer_api_key_ref
            used_secret_refs.add(transfer_api_key_ref)
        if transfer_api_secret:
            transfer_api_secret_ref = main_account_secret_ref(main_id, "transfer_api_secret")
            next_store[transfer_api_secret_ref] = transfer_api_secret
            rendered_main_account["transfer_api_secret_secret_ref"] = transfer_api_secret_ref
            used_secret_refs.add(transfer_api_secret_ref)

        children = raw_main_account.get("children")
        if not isinstance(children, list):
            raise AccountImportError(f"Main account {main_id} must define children")
        rendered_children: list[dict[str, Any]] = []
        for raw_child in children:
            if not isinstance(raw_child, dict):
                raise AccountImportError(f"Children for {main_id} must be objects")
            account_id = str(raw_child.get("account_id") or "").strip().lower()
            if not account_id:
                raise AccountImportError(f"Child account under {main_id} must define account_id")
            api_key_ref = child_account_secret_ref(main_id, account_id, "api_key")
            api_secret_ref = child_account_secret_ref(main_id, account_id, "api_secret")
            api_key = str(raw_child.get("api_key") or "").strip()
            api_secret = str(raw_child.get("api_secret") or "").strip()
            if not api_key or not api_secret:
                raise AccountImportError(f"Child account {main_id}.{account_id} must define api_key and api_secret")
            next_store[api_key_ref] = api_key
            next_store[api_secret_ref] = api_secret
            used_secret_refs.update({api_key_ref, api_secret_ref})
            rendered_children.append(
                {
                    "account_id": account_id,
                    "name": str(raw_child.get("name") or "").strip(),
                    "api_key_secret_ref": api_key_ref,
                    "api_secret_secret_ref": api_secret_ref,
                    "uid": str(raw_child.get("uid") or "").strip(),
                    "use_testnet": bool(raw_child.get("use_testnet")),
                    "rest_base_url": str(raw_child.get("rest_base_url") or "").strip(),
                    "ws_base_url": str(raw_child.get("ws_base_url") or "").strip(),
                }
            )

        rendered_main_account["children"] = rendered_children
        rendered_main_accounts.append(rendered_main_account)

    return {"main_accounts": rendered_main_accounts}, used_secret_refs, next_store


def collect_monitor_account_secret_refs(payload: dict[str, Any]) -> set[str]:
    refs: set[str] = set()
    main_accounts = payload.get("main_accounts") if isinstance(payload, dict) else None
    if not isinstance(main_accounts, list):
        return refs
    for main_account in main_accounts:
        if not isinstance(main_account, dict):
            continue
        for field_name in ("transfer_api_key_secret_ref", "transfer_api_secret_secret_ref"):
            value = str(main_account.get(field_name) or "").strip()
            if value:
                refs.add(value)
        children = main_account.get("children")
        if not isinstance(children, list):
            continue
        for child in children:
            if not isinstance(child, dict):
                continue
            for field_name in ("api_key_secret_ref", "api_secret_secret_ref"):
                value = str(child.get(field_name) or "").strip()
                if value:
                    refs.add(value)
    return refs


def delete_unused_monitor_account_secret_refs(
    *,
    secret_provider: EncryptedFileSecretProvider,
    previous_payload: dict[str, Any] | None,
    retained_secret_refs: set[str],
) -> None:
    if not previous_payload:
        return
    previous_refs = collect_monitor_account_secret_refs(previous_payload)
    for secret_ref in sorted(previous_refs - set(retained_secret_refs)):
        secret_provider.delete_secret(secret_ref)


def apply_settings_secret_updates(
    *,
    settings_updates: dict[str, str],
    current_store: dict[str, str],
    access_control_payload: dict[str, Any],
    env_content: str,
    master_key_file: str = "",
    setting_specs: dict[str, SettingSpec] | None = None,
) -> tuple[list[str], dict[str, Any], str, dict[str, str]]:
    if not settings_updates:
        normalized_env_content = env_content if env_content.endswith("\n") else f"{env_content}\n"
        return [], dict(access_control_payload), normalized_env_content, dict(current_store)

    specs = setting_specs or DEFAULT_SETTING_SPECS
    next_store = {str(key): str(value) for key, value in dict(current_store).items()}
    next_access_control_payload = dict(access_control_payload)
    next_access_control_payload.setdefault("enabled", False)
    next_access_control_payload.setdefault("whitelist_ips", [])
    next_access_control_payload["allow_plaintext_secrets"] = False
    next_access_control_payload.setdefault("cookie_secure_mode", "auto")
    env_lines = str(env_content or "").splitlines()
    env_values: dict[str, str] = {}
    for raw_line in env_lines:
        if not raw_line or raw_line.lstrip().startswith("#") or "=" not in raw_line:
            continue
        key, value = raw_line.split("=", 1)
        env_values[key.strip()] = value

    for spec in specs.values():
        if spec.target == "access_control":
            existing_ref = str(next_access_control_payload.get(spec.ref_field) or "").strip()
            existing_plaintext = str(next_access_control_payload.get(spec.plaintext_field) or "").strip()
            if existing_plaintext:
                next_store[spec.secret_ref] = existing_plaintext
                next_access_control_payload[spec.ref_field] = spec.secret_ref
                next_access_control_payload.pop(spec.plaintext_field, None)
            elif existing_ref and existing_ref in next_store:
                next_access_control_payload[spec.ref_field] = existing_ref
        elif spec.target == "env":
            existing_plaintext = str(env_values.get(spec.plaintext_field) or "").strip()
            existing_ref = str(env_values.get(spec.env_var) or "").strip()
            if existing_plaintext:
                next_store[spec.secret_ref] = existing_plaintext
                env_values.pop(spec.plaintext_field, None)
                env_values[spec.env_var] = spec.secret_ref
            elif existing_ref and existing_ref in next_store:
                env_values[spec.env_var] = existing_ref

    updated_keys: list[str] = []
    for key, value in settings_updates.items():
        spec = specs[key]
        next_store[spec.secret_ref] = value
        if spec.target == "access_control":
            next_access_control_payload["allow_plaintext_secrets"] = False
            next_access_control_payload[spec.ref_field] = spec.secret_ref
            next_access_control_payload.pop(spec.plaintext_field, None)
        elif spec.target == "env":
            env_values[spec.env_var] = spec.secret_ref
            env_values.pop(spec.plaintext_field, None)
        updated_keys.append(key)

    normalized_master_key_file = str(master_key_file or "").strip()
    if normalized_master_key_file:
        env_values["MONITOR_MASTER_KEY_FILE"] = normalized_master_key_file

    rendered_env_lines = _rewrite_env_lines(env_lines, env_values)
    return sorted(updated_keys), next_access_control_payload, rendered_env_lines, next_store


def build_accounts_excel_template(
    *,
    setting_specs: dict[str, SettingSpec] | None = None,
) -> bytes:
    specs = setting_specs or DEFAULT_SETTING_SPECS

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "accounts"
    worksheet.append(list(SUPPORTED_COLUMNS))
    worksheet.freeze_panes = "A2"
    for column_name, width in {
        "A": 16,
        "B": 18,
        "C": 14,
        "D": 18,
        "E": 32,
        "F": 32,
        "G": 16,
        "H": 14,
        "I": 24,
        "J": 24,
    }.items():
        worksheet.column_dimensions[column_name].width = width

    settings_sheet = workbook.create_sheet(title="settings")
    settings_sheet.append(list(SUPPORTED_SETTINGS_COLUMNS))
    for key in sorted(specs):
        settings_sheet.append([key, "", "leave blank to keep the existing encrypted value"])
    settings_sheet.freeze_panes = "A2"
    settings_sheet.column_dimensions["A"].width = 32
    settings_sheet.column_dimensions["B"].width = 42
    settings_sheet.column_dimensions["C"].width = 56

    guide = workbook.create_sheet(title="guide")
    guide.append(("section", "content"))
    guide.append(("template_version", TEMPLATE_VERSION))
    guide.append(("purpose", "Use the accounts sheet to import Binance monitor accounts in replace_all mode. Leave it header-only when you only want to update settings."))
    guide.append(("main-row", "Set account_id to main for the main transfer row; api_key/api_secret map to transfer_api_key/transfer_api_secret and uid is required."))
    guide.append(("child-row", "Each child account uses its own row with account_id, api_key, api_secret, and optional uid/testnet/base URLs."))
    guide.append(("settings", "Use the settings sheet to update supported global secrets incrementally. Filled values overwrite the encrypted value; blanks keep the existing value."))
    guide.append(("validation", "main_id/account_id must be normalized IDs; the same main_id must keep a consistent main_name; duplicate child account_id under one main_id is not allowed."))
    guide.append(("security", "API keys, passwords, and tokens are accepted in plaintext during import, then written into the encrypted secret store and replaced with *_secret_ref in project config."))
    guide.append(("cleanup", "After import succeeds, delete the local Excel file that still contains plaintext sensitive values."))
    guide.freeze_panes = "A2"
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 120

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _parse_template_version(workbook) -> str:
    if "guide" not in workbook.sheetnames:
        return "legacy"
    guide = workbook["guide"]
    for row in guide.iter_rows(values_only=True):
        if _parse_optional_text(row[0] if row else "").lower() == "template_version":
            return _parse_optional_text(row[1] if len(row) > 1 else "")
    return "legacy"


def _parse_accounts_worksheet(worksheet: Worksheet, *, filename: str) -> tuple[dict[str, Any], ImportResult]:
    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        raise AccountImportError("Excel file must contain a header row")

    header_map = _build_header_map(header_row)
    grouped_rows: dict[str, dict[str, Any]] = {}

    for row_index, row in enumerate(rows, start=2):
        row_data = _extract_row_data(row, header_map)
        if _is_blank_row(row_data):
            continue

        main_id = _parse_required_id(row_data["main_id"], field_name="main_id", row_index=row_index)
        main_name = _parse_required_text(row_data["main_name"], field_name="main_name", row_index=row_index)
        account_id = _parse_required_id(row_data["account_id"], field_name="account_id", row_index=row_index)
        row_name = _parse_required_text(row_data["name"], field_name="name", row_index=row_index)
        api_key = _parse_required_text(row_data["api_key"], field_name="api_key", row_index=row_index)
        api_secret = _parse_required_text(row_data["api_secret"], field_name="api_secret", row_index=row_index)
        uid = _parse_optional_text(row_data["uid"])
        use_testnet = _parse_optional_bool(row_data["use_testnet"], field_name="use_testnet", row_index=row_index)
        rest_base_url = _parse_optional_text(row_data["rest_base_url"])
        ws_base_url = _parse_optional_text(row_data["ws_base_url"])

        grouped = grouped_rows.setdefault(
            main_id,
            {
                "main_id": main_id,
                "name": main_name,
                "transfer_api_key": "",
                "transfer_api_secret": "",
                "transfer_uid": "",
                "children": [],
                "_child_ids": set(),
                "_has_main_row": False,
            },
        )
        if grouped["name"] != main_name:
            raise AccountImportError(f"Row {row_index}: main_name must stay consistent for main_id {main_id}")

        if account_id == "main":
            if grouped["_has_main_row"]:
                raise AccountImportError(f"Row {row_index}: duplicate reserved main row under main_id {main_id}")
            if not uid:
                raise AccountImportError(f"Row {row_index}: uid is required for reserved main row")
            grouped["transfer_api_key"] = api_key
            grouped["transfer_api_secret"] = api_secret
            grouped["transfer_uid"] = uid
            grouped["_has_main_row"] = True
            continue

        if account_id in grouped["_child_ids"]:
            raise AccountImportError(f"Row {row_index}: duplicate account_id {account_id} under main_id {main_id}")

        grouped["children"].append(
            {
                "account_id": account_id,
                "name": row_name,
                "api_key": api_key,
                "api_secret": api_secret,
                "uid": uid,
                "use_testnet": use_testnet,
                "rest_base_url": rest_base_url,
                "ws_base_url": ws_base_url,
            }
        )
        grouped["_child_ids"].add(account_id)

    payload = {
        "main_accounts": [
            {
                "main_id": grouped["main_id"],
                "name": grouped["name"],
                "transfer_api_key": grouped["transfer_api_key"],
                "transfer_api_secret": grouped["transfer_api_secret"],
                "transfer_uid": grouped["transfer_uid"],
                "children": grouped["children"],
            }
            for _, grouped in sorted(grouped_rows.items(), key=lambda item: item[0])
        ]
    }
    parse_monitor_accounts_payload(payload, allow_plaintext_secrets=True)
    result = ImportResult(
        file_name=filename,
        main_account_count=len(payload["main_accounts"]),
        account_count=sum(len(main_account["children"]) for main_account in payload["main_accounts"]),
    )
    return payload, result


def _parse_settings_worksheet(
    worksheet: Worksheet,
    *,
    setting_specs: dict[str, SettingSpec],
) -> dict[str, str]:
    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        return {}
    header_map = _build_settings_header_map(header_row)
    updates: dict[str, str] = {}
    for row_index, row in enumerate(rows, start=2):
        row_data = _extract_settings_row_data(row, header_map)
        if _is_blank_row(row_data):
            continue
        key = _parse_required_text(row_data["key"], field_name="key", row_index=row_index)
        if key not in setting_specs:
            raise AccountImportError(f"Row {row_index}: unsupported settings key {key}")
        value = _parse_optional_text(row_data["value"])
        if not value:
            continue
        if key in updates:
            raise AccountImportError(f"Row {row_index}: duplicate settings key {key}")
        updates[key] = value
    return updates


def _build_header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, value in enumerate(header_row):
        name = _parse_optional_text(value).lower()
        if not name:
            continue
        if name in header_map:
            raise AccountImportError(f"Duplicate header: {name}")
        header_map[name] = index

    missing = [column for column in REQUIRED_COLUMNS if column not in header_map]
    if missing:
        raise AccountImportError(f"Missing required columns: {', '.join(missing)}")
    return header_map


def _build_settings_header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, value in enumerate(header_row):
        name = _parse_optional_text(value).lower()
        if not name:
            continue
        if name in header_map:
            raise AccountImportError(f"Duplicate header: {name}")
        header_map[name] = index
    missing = [column for column in SETTINGS_REQUIRED_COLUMNS if column not in header_map]
    if missing:
        raise AccountImportError(f"Missing required settings columns: {', '.join(missing)}")
    return header_map


def _extract_row_data(row: tuple[Any, ...], header_map: dict[str, int]) -> dict[str, Any]:
    data: dict[str, Any] = {}
    for column in SUPPORTED_COLUMNS:
        index = header_map.get(column)
        data[column] = row[index] if index is not None and index < len(row) else None
    return data


def _extract_settings_row_data(row: tuple[Any, ...], header_map: dict[str, int]) -> dict[str, Any]:
    data: dict[str, Any] = {}
    for column in SUPPORTED_SETTINGS_COLUMNS:
        index = header_map.get(column)
        data[column] = row[index] if index is not None and index < len(row) else None
    return data


def _is_blank_row(row_data: dict[str, Any]) -> bool:
    return all(_parse_optional_text(value) == "" for value in row_data.values())


def _parse_required_id(value: Any, *, field_name: str, row_index: int) -> str:
    try:
        return normalize_monitor_id(value, field_name=field_name)
    except ValueError as exc:
        raise AccountImportError(f"Row {row_index}: {exc}") from exc


def _parse_required_text(value: Any, *, field_name: str, row_index: int) -> str:
    normalized = _parse_optional_text(value)
    if not normalized:
        raise AccountImportError(f"Row {row_index}: {field_name} is required")
    return normalized


def _parse_optional_text(value: Any) -> str:
    return str(value or "").strip()


def _parse_optional_bool(value: Any, *, field_name: str, row_index: int) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if not normalized:
        return False
    if normalized in TRUE_VALUES:
        return True
    if normalized in FALSE_VALUES:
        return False
    raise AccountImportError(f"Row {row_index}: {field_name} must be one of true/false/1/0/yes/no/on/off")


def _write_text_atomic(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            dir=path.parent,
            delete=False,
            prefix=f"{path.stem}-",
            suffix=".tmp",
        ) as handle:
            handle.write(content)
            temp_path = Path(handle.name)
        os.replace(temp_path, path)
    except OSError:
        if temp_path is not None and temp_path.exists():
            temp_path.unlink(missing_ok=True)
        raise


def _rewrite_env_lines(existing_lines: list[str], values: dict[str, str]) -> str:
    rendered_lines: list[str] = []
    seen_keys: set[str] = set()
    managed_keys = {"TG_BOT_TOKEN", "TG_CHAT_ID", "TG_BOT_TOKEN_SECRET_REF", "TG_CHAT_ID_SECRET_REF", "MONITOR_MASTER_KEY_FILE"}
    for raw_line in existing_lines:
        if not raw_line or raw_line.lstrip().startswith("#") or "=" not in raw_line:
            rendered_lines.append(raw_line)
            continue
        key, _ = raw_line.split("=", 1)
        normalized = key.strip()
        if normalized in {"TG_BOT_TOKEN", "TG_CHAT_ID"}:
            seen_keys.add(normalized)
            continue
        if normalized in values and normalized in managed_keys:
            rendered_lines.append(f"{normalized}={values[normalized]}")
            seen_keys.add(normalized)
            continue
        rendered_lines.append(raw_line)
        seen_keys.add(normalized)

    for key in ("TG_BOT_TOKEN_SECRET_REF", "TG_CHAT_ID_SECRET_REF", "MONITOR_MASTER_KEY_FILE"):
        if key in values and key not in seen_keys:
            rendered_lines.append(f"{key}={values[key]}")
    return "\n".join(rendered_lines).rstrip() + "\n"
