from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook

from monitor_app.account_import import AccountImportError, TEMPLATE_VERSION, parse_accounts_excel


def test_parse_accounts_excel_builds_hierarchical_payload_with_transfer_config() -> None:
    content = _build_workbook_bytes(
        account_rows=[
            ["main_id", "main_name", "account_id", "name", "api_key", "api_secret", "uid", "use_testnet"],
            ["group_a", "Group A", "main", "Main Transfer", "mk1", "ms1", "123456789", ""],
            ["group_a", "Group A", "sub1", "Sub One", "k1", "s1", "223456789", "true"],
            ["group_a", "Group A", "sub2", "Sub Two", "k2", "s2", "", ""],
            ["group_b", "Group B", "sub1", "Sub Three", "k3", "s3", "323456789", "false"],
        ]
    )

    parsed = parse_accounts_excel(content, filename="accounts.xlsx")

    assert parsed.template_version == TEMPLATE_VERSION
    assert parsed.import_result.file_name == "accounts.xlsx"
    assert parsed.import_result.main_account_count == 2
    assert parsed.import_result.account_count == 3
    assert parsed.import_result.updated_settings_keys == ()
    assert parsed.payload == {
        "main_accounts": [
            {
                "main_id": "group_a",
                "name": "Group A",
                "transfer_api_key": "mk1",
                "transfer_api_secret": "ms1",
                "transfer_uid": "123456789",
                "children": [
                    {
                        "account_id": "sub1",
                        "name": "Sub One",
                        "api_key": "k1",
                        "api_secret": "s1",
                        "uid": "223456789",
                        "use_testnet": True,
                        "rest_base_url": "",
                        "ws_base_url": "",
                    },
                    {
                        "account_id": "sub2",
                        "name": "Sub Two",
                        "api_key": "k2",
                        "api_secret": "s2",
                        "uid": "",
                        "use_testnet": False,
                        "rest_base_url": "",
                        "ws_base_url": "",
                    },
                ],
            },
            {
                "main_id": "group_b",
                "name": "Group B",
                "transfer_api_key": "",
                "transfer_api_secret": "",
                "transfer_uid": "",
                "children": [
                    {
                        "account_id": "sub1",
                        "name": "Sub Three",
                        "api_key": "k3",
                        "api_secret": "s3",
                        "uid": "323456789",
                        "use_testnet": False,
                        "rest_base_url": "",
                        "ws_base_url": "",
                    }
                ],
            },
        ]
    }


def test_parse_accounts_excel_collects_supported_settings_updates() -> None:
    content = _build_workbook_bytes(
        account_rows=[
            ["main_id", "main_name", "account_id", "name", "api_key", "api_secret"],
            ["group_a", "Group A", "sub1", "Sub One", "k1", "s1"],
        ],
        settings_rows=[
            ["key", "value", "notes"],
            ["telegram.bot_token", "bot-token", "replace tg token"],
            ["access_control.admin_password", "admin-pass", "replace admin password"],
            ["telegram.chat_id", "", "blank should keep old value"],
        ],
    )

    parsed = parse_accounts_excel(content, filename="accounts.xlsx")

    assert parsed.settings_updates == {
        "telegram.bot_token": "bot-token",
        "access_control.admin_password": "admin-pass",
    }
    assert parsed.import_result.updated_settings_keys == (
        "access_control.admin_password",
        "telegram.bot_token",
    )


def test_parse_accounts_excel_rejects_settings_sheet_without_template_version() -> None:
    content = _build_workbook_bytes(
        account_rows=[
            ["main_id", "main_name", "account_id", "name", "api_key", "api_secret"],
            ["group_a", "Group A", "sub1", "Sub One", "k1", "s1"],
        ],
        settings_rows=[
            ["key", "value"],
            ["telegram.bot_token", "bot-token"],
        ],
        include_guide=False,
    )

    with pytest.raises(AccountImportError, match="settings worksheet requires a supported template_version"):
        parse_accounts_excel(content, filename="accounts.xlsx")


def test_parse_accounts_excel_allows_settings_only_template() -> None:
    content = _build_workbook_bytes(
        account_rows=[
            ["main_id", "main_name", "account_id", "name", "api_key", "api_secret"],
        ],
        settings_rows=[
            ["key", "value", "notes"],
            ["access_control.guest_password", "guest-pass-next", "rotate guest password"],
        ],
    )

    parsed = parse_accounts_excel(content, filename="accounts.xlsx")

    assert parsed.payload == {"main_accounts": []}
    assert parsed.import_result.main_account_count == 0
    assert parsed.import_result.account_count == 0
    assert parsed.import_result.mode == "settings_only"
    assert parsed.settings_updates == {"access_control.guest_password": "guest-pass-next"}
    assert parsed.import_result.updated_settings_keys == ("access_control.guest_password",)


def test_parse_accounts_excel_rejects_empty_accounts_and_settings() -> None:
    content = _build_workbook_bytes(
        account_rows=[
            ["main_id", "main_name", "account_id", "name", "api_key", "api_secret"],
        ],
        settings_rows=[
            ["key", "value", "notes"],
        ],
    )

    with pytest.raises(
        AccountImportError,
        match="Excel workbook must include at least one account row or one settings update",
    ):
        parse_accounts_excel(content, filename="accounts.xlsx")


def test_parse_accounts_excel_rejects_unsupported_settings_key() -> None:
    content = _build_workbook_bytes(
        account_rows=[
            ["main_id", "main_name", "account_id", "name", "api_key", "api_secret"],
            ["group_a", "Group A", "sub1", "Sub One", "k1", "s1"],
        ],
        settings_rows=[
            ["key", "value"],
            ["some.random.secret", "value"],
        ],
    )

    with pytest.raises(AccountImportError, match=r"Row 2: unsupported settings key some.random.secret"):
        parse_accounts_excel(content, filename="accounts.xlsx")


def test_parse_accounts_excel_rejects_missing_required_headers() -> None:
    content = _build_workbook_bytes(
        account_rows=[
            ["main_id", "main_name", "account_id", "name", "api_key"],
            ["group_a", "Group A", "sub1", "Sub One", "k1"],
        ]
    )

    with pytest.raises(AccountImportError, match="Missing required columns: api_secret"):
        parse_accounts_excel(content, filename="accounts.xlsx")


def test_parse_accounts_excel_rejects_blank_required_value() -> None:
    content = _build_workbook_bytes(
        account_rows=[
            ["main_id", "main_name", "account_id", "name", "api_key", "api_secret"],
            ["group_a", "Group A", "sub1", "Sub One", "", "s1"],
        ]
    )

    with pytest.raises(AccountImportError, match=r"Row 2: api_key is required"):
        parse_accounts_excel(content, filename="accounts.xlsx")


def test_parse_accounts_excel_rejects_invalid_account_id() -> None:
    content = _build_workbook_bytes(
        account_rows=[
            ["main_id", "main_name", "account_id", "name", "api_key", "api_secret"],
            ["group_a", "Group A", "Sub 1", "Sub One", "k1", "s1"],
        ]
    )

    with pytest.raises(AccountImportError, match=r"Row 2: account_id must match"):
        parse_accounts_excel(content, filename="accounts.xlsx")


def test_parse_accounts_excel_rejects_duplicate_account_id_within_group() -> None:
    content = _build_workbook_bytes(
        account_rows=[
            ["main_id", "main_name", "account_id", "name", "api_key", "api_secret"],
            ["group_a", "Group A", "sub1", "Sub One", "k1", "s1"],
            ["group_a", "Group A", "sub1", "Sub Two", "k2", "s2"],
        ]
    )

    with pytest.raises(AccountImportError, match=r"Row 3: duplicate account_id sub1 under main_id group_a"):
        parse_accounts_excel(content, filename="accounts.xlsx")


def test_parse_accounts_excel_rejects_duplicate_reserved_main_row() -> None:
    content = _build_workbook_bytes(
        account_rows=[
            ["main_id", "main_name", "account_id", "name", "api_key", "api_secret", "uid"],
            ["group_a", "Group A", "main", "Main Transfer", "mk1", "ms1", "123456789"],
            ["group_a", "Group A", "main", "Main Transfer Again", "mk2", "ms2", "223456789"],
            ["group_a", "Group A", "sub1", "Sub One", "k1", "s1", "323456789"],
        ]
    )

    with pytest.raises(AccountImportError, match=r"Row 3: duplicate reserved main row under main_id group_a"):
        parse_accounts_excel(content, filename="accounts.xlsx")


def test_parse_accounts_excel_requires_uid_for_reserved_main_row() -> None:
    content = _build_workbook_bytes(
        account_rows=[
            ["main_id", "main_name", "account_id", "name", "api_key", "api_secret", "uid"],
            ["group_a", "Group A", "main", "Main Transfer", "mk1", "ms1", ""],
            ["group_a", "Group A", "sub1", "Sub One", "k1", "s1", "323456789"],
        ]
    )

    with pytest.raises(AccountImportError, match=r"Row 2: uid is required for reserved main row"):
        parse_accounts_excel(content, filename="accounts.xlsx")


def test_parse_accounts_excel_rejects_inconsistent_group_name() -> None:
    content = _build_workbook_bytes(
        account_rows=[
            ["main_id", "main_name", "account_id", "name", "api_key", "api_secret"],
            ["group_a", "Group A", "sub1", "Sub One", "k1", "s1"],
            ["group_a", "Group X", "sub2", "Sub Two", "k2", "s2"],
        ]
    )

    with pytest.raises(AccountImportError, match=r"Row 3: main_name must stay consistent for main_id group_a"):
        parse_accounts_excel(content, filename="accounts.xlsx")


def test_parse_accounts_excel_defaults_optional_columns() -> None:
    content = _build_workbook_bytes(
        account_rows=[
            ["main_id", "main_name", "account_id", "name", "api_key", "api_secret"],
            ["group_a", "Group A", "sub1", "Sub One", "k1", "s1"],
            [None, None, None, None, None, None],
        ]
    )

    parsed = parse_accounts_excel(content, filename="accounts.xlsx")

    child = parsed.payload["main_accounts"][0]["children"][0]
    assert child["uid"] == ""
    assert child["use_testnet"] is False
    assert child["rest_base_url"] == ""
    assert child["ws_base_url"] == ""


def _build_workbook_bytes(
    *,
    account_rows: list[list[object]],
    settings_rows: list[list[object]] | None = None,
    include_guide: bool = True,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "accounts"
    for row in account_rows:
        worksheet.append(row)

    if settings_rows is not None:
        settings_sheet = workbook.create_sheet("settings")
        for row in settings_rows:
            settings_sheet.append(row)

    if include_guide:
        guide = workbook.create_sheet("guide")
        guide.append(["section", "content"])
        guide.append(["template_version", TEMPLATE_VERSION])

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()
