from __future__ import annotations

import asyncio
import json
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook
import pytest

import monitor_app.api as api_module
from monitor_app.access_control.service import AUTH_COOKIE_NAME, AccessControlService
from monitor_app.account_monitor import AccountMonitorController
from monitor_app.account_import import TEMPLATE_VERSION
from monitor_app.config import MonitorAccountConfig, Settings
from monitor_app.i18n import (
    excel_import_refresh_failed_message,
    excel_import_refresh_success_message,
    excel_import_settings_success_message,
)
from monitor_app.secrets import EncryptedFileSecretProvider, create_master_key


class FakeImportGateway:
    def __init__(self, account: MonitorAccountConfig) -> None:
        self.account = account

    async def get_unified_account_snapshot(
        self,
        *,
        history_window_days: int = 7,
        income_limit: int = 100,
        interest_limit: int = 100,
        previous_snapshot: dict | None = None,
        mark_price_provider=None,
        refresh_id: str | None = None,
        refresh_reason: str | None = None,
    ) -> dict:
        return {
            "status": "ok",
            "source": "papi",
            "account_id": self.account.account_id,
            "account_name": self.account.display_name,
            "main_account_id": self.account.main_account_id,
            "main_account_name": self.account.main_account_name,
            "child_account_id": self.account.child_account_id,
            "child_account_name": self.account.child_account_name,
            "account_status": "NORMAL",
            "updated_at": "2026-03-27T00:00:00+00:00",
            "message": "ok",
            "totals": {
                "equity": Decimal("100"),
                "margin": Decimal("10"),
                "available_balance": Decimal("90"),
                "unrealized_pnl": Decimal("5"),
                "total_income": Decimal("0"),
                "total_commission": Decimal("0"),
                "total_distribution": Decimal("0"),
                "distribution_apy_7d": Decimal("0"),
                "total_interest": Decimal("0"),
            },
            "positions": [],
            "assets": [],
            "spot_assets": [],
            "income_summary": {
                "window_days": history_window_days,
                "records": 0,
                "total_income": Decimal("0"),
                "total_commission": Decimal("0"),
                "by_type": {},
                "by_asset": {},
            },
            "distribution_summary": {
                "window_days": history_window_days,
                "records": 0,
                "total_distribution": Decimal("0"),
                "by_type": {},
                "by_asset": {},
            },
            "distribution_profit_summary": {
                "today": {"label": "today", "amount": Decimal("0"), "rate": Decimal("0"), "start_at": None, "complete": True},
                "week": {"label": "week", "amount": Decimal("0"), "rate": Decimal("0"), "start_at": None, "complete": True},
                "month": {"label": "month", "amount": Decimal("0"), "rate": Decimal("0"), "start_at": None, "complete": True},
                "year": {"label": "year", "amount": Decimal("0"), "rate": Decimal("0"), "start_at": None, "complete": True},
                "all": {"label": "all", "amount": Decimal("0"), "rate": Decimal("0"), "start_at": None, "complete": True},
                "backfill_complete": True,
            },
            "interest_summary": {
                "window_days": history_window_days,
                "records": 0,
                "margin_interest_total": Decimal("0"),
                "negative_balance_interest_total": Decimal("0"),
                "total_interest": Decimal("0"),
            },
            "section_errors": {},
            "diagnostics": {
                "refresh_id": refresh_id,
                "timings": {"gateway_total_ms": 1},
                "fallback_sections": [],
            },
        }

    async def close(self) -> None:
        return None


class FailingImportGateway(FakeImportGateway):
    async def get_unified_account_snapshot(
        self,
        *,
        history_window_days: int = 7,
        income_limit: int = 100,
        interest_limit: int = 100,
        previous_snapshot: dict | None = None,
        mark_price_provider=None,
        refresh_id: str | None = None,
        refresh_reason: str | None = None,
    ) -> dict:
        raise RuntimeError("refresh failed")


class ReloadThenFailController(AccountMonitorController):
    async def reload_accounts(self) -> dict[str, object]:
        await super().reload_accounts()
        raise RuntimeError("reload failed after state change")


def _build_disabled_access_control(tmp_path: Path, settings: Settings) -> AccessControlService:
    config_path = tmp_path / "access_control.json"
    audit_db_path = tmp_path / "access_audit.db"
    config_path.parent.mkdir(parents=True, exist_ok=True)
    config_path.write_text(
        json.dumps(
            {
                "enabled": False,
                "whitelist_ips": [],
                "allow_plaintext_secrets": True,
                "cookie_secure_mode": "auto",
                "guest_password": "guest-pass",
                "admin_password": "admin-pass",
                "session_secret": "dev-session-secret-1234567890",
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return AccessControlService(settings, config_path=config_path, audit_db_path=audit_db_path)


def _build_enabled_access_control(tmp_path: Path, settings: Settings) -> AccessControlService:
    config_path = settings.access_control_config_file
    audit_db_path = tmp_path / "access_audit.db"
    config_path.parent.mkdir(parents=True, exist_ok=True)
    config_path.write_text(
        json.dumps(
            {
                "enabled": True,
                "whitelist_ips": [],
                "allow_plaintext_secrets": True,
                "cookie_secure_mode": "auto",
                "guest_password": "guest-pass",
                "admin_password": "admin-pass",
                "session_secret": "dev-session-secret-1234567890",
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return AccessControlService(settings, config_path=config_path, audit_db_path=audit_db_path)


def test_import_excel_endpoint_replaces_accounts_and_refreshes(monkeypatch, tmp_path: Path) -> None:
    master_key = create_master_key()
    env_path = tmp_path / ".env"
    env_path.write_text("TG_ENABLED=true\n", encoding="utf-8")
    settings = Settings(
        _env_file=None,
        monitor_accounts_file=tmp_path / "config" / "binance_monitor_accounts.json",
        access_control_config_file=tmp_path / "access_control.json",
        secrets_file=tmp_path / "config" / "secrets.enc.json",
        env_file_path=env_path,
        monitor_master_key=master_key,
        monitor_refresh_interval_ms=999999,
        monitor_history_window_days=3,
        monitor_history_db_path=tmp_path / "history.db",
    )
    monkeypatch.setattr(api_module, "settings", settings)

    with TestClient(api_module.app) as client:
        client.app.state.access_control = _build_disabled_access_control(tmp_path, settings)
        controller = AccountMonitorController(settings, gateway_factory=lambda account: FakeImportGateway(account))
        client.app.state.monitor = controller
        try:
            response = client.post(
                "/api/config/import/excel",
                files={"file": ("accounts.xlsx", _build_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        finally:
            asyncio.run(controller.close())

    assert response.status_code == 200
    payload = response.json()
    assert payload["message"] == excel_import_refresh_success_message()
    assert payload["import_result"]["file_name"] == "accounts.xlsx"
    assert payload["import_result"]["main_account_count"] == 2
    assert payload["import_result"]["account_count"] == 3
    assert payload["import_result"]["updated_settings_keys"] == []
    assert payload["import_result"]["updated_secret_refs"] == [
        "accounts.group_a.sub1.api_key",
        "accounts.group_a.sub1.api_secret",
        "accounts.group_a.sub2.api_key",
        "accounts.group_a.sub2.api_secret",
        "accounts.group_b.sub1.api_key",
        "accounts.group_b.sub1.api_secret",
        "main_accounts.group_a.transfer_api_key",
        "main_accounts.group_a.transfer_api_secret",
    ]
    assert payload["import_result"]["mode"] == "replace_all"
    assert payload["template_version"] == TEMPLATE_VERSION
    assert "加密仓库" in payload["security_notice"]
    assert payload["refresh_result"]["success"] is True
    assert payload["service"]["account_ids"] == ["group_a.sub1", "group_a.sub2", "group_b.sub1"]
    assert payload["service"]["main_account_ids"] == ["group_a", "group_b"]
    persisted = json.loads(settings.monitor_accounts_file.read_text(encoding="utf-8"))
    assert len(persisted["main_accounts"]) == 2
    assert persisted["main_accounts"][0]["main_id"] == "group_a"
    assert persisted["main_accounts"][0]["transfer_api_key_secret_ref"] == "main_accounts.group_a.transfer_api_key"
    assert persisted["main_accounts"][0]["transfer_api_secret_secret_ref"] == "main_accounts.group_a.transfer_api_secret"
    assert persisted["main_accounts"][0]["transfer_uid"] == "123456789"
    assert "transfer_api_key" not in persisted["main_accounts"][0]
    assert persisted["main_accounts"][0]["children"][0]["uid"] == "223456789"
    assert persisted["main_accounts"][0]["children"][0]["api_key_secret_ref"] == "accounts.group_a.sub1.api_key"
    assert persisted["main_accounts"][0]["children"][0]["api_secret_secret_ref"] == "accounts.group_a.sub1.api_secret"
    assert "transfer_api_key_secret_ref" not in persisted["main_accounts"][1]
    provider = EncryptedFileSecretProvider(settings.secrets_file, master_key=master_key)
    assert provider.get_secret("main_accounts.group_a.transfer_api_key") == "mk1"
    assert provider.get_secret("main_accounts.group_a.transfer_api_secret") == "ms1"
    assert provider.get_secret("accounts.group_a.sub1.api_key") == "k1"
    assert provider.get_secret("accounts.group_b.sub1.api_secret") == "s3"


def test_import_excel_endpoint_updates_supported_settings_as_secret_refs(monkeypatch, tmp_path: Path) -> None:
    master_key = create_master_key()
    env_path = tmp_path / ".env"
    env_path.write_text("TG_ENABLED=true\n", encoding="utf-8")
    settings = Settings(
        _env_file=None,
        monitor_accounts_file=tmp_path / "config" / "binance_monitor_accounts.json",
        access_control_config_file=tmp_path / "access_control.json",
        secrets_file=tmp_path / "config" / "secrets.enc.json",
        env_file_path=env_path,
        monitor_master_key=master_key,
        monitor_refresh_interval_ms=999999,
        monitor_history_window_days=3,
        monitor_history_db_path=tmp_path / "history.db",
    )
    monkeypatch.setattr(api_module, "settings", settings)

    with TestClient(api_module.app) as client:
        client.app.state.access_control = _build_disabled_access_control(tmp_path, settings)
        controller = AccountMonitorController(settings, gateway_factory=lambda account: FakeImportGateway(account))
        client.app.state.monitor = controller
        try:
            response = client.post(
                "/api/config/import/excel",
                files={
                    "file": (
                        "accounts.xlsx",
                        _build_workbook_bytes(
                            settings_rows=[
                                ["key", "value", "notes"],
                                ["telegram.bot_token", "bot-token", "update token"],
                                ["telegram.chat_id", "chat-id", "update chat id"],
                                ["access_control.admin_password", "admin-pass-next", "rotate admin password"],
                            ],
                        ),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        finally:
            asyncio.run(controller.close())

    assert response.status_code == 200
    payload = response.json()
    assert payload["import_result"]["updated_settings_keys"] == [
        "access_control.admin_password",
        "telegram.bot_token",
        "telegram.chat_id",
    ]
    assert payload["import_result"]["updated_secret_refs"] == [
        "access_control.admin_password",
        "access_control.guest_password",
        "access_control.session_secret",
        "accounts.group_a.sub1.api_key",
        "accounts.group_a.sub1.api_secret",
        "accounts.group_a.sub2.api_key",
        "accounts.group_a.sub2.api_secret",
        "accounts.group_b.sub1.api_key",
        "accounts.group_b.sub1.api_secret",
        "main_accounts.group_a.transfer_api_key",
        "main_accounts.group_a.transfer_api_secret",
        "telegram.bot_token",
        "telegram.chat_id",
    ]
    access_control_payload = json.loads(settings.access_control_config_file.read_text(encoding="utf-8"))
    assert access_control_payload["allow_plaintext_secrets"] is False
    assert access_control_payload["guest_password_secret_ref"] == "access_control.guest_password"
    assert access_control_payload["admin_password_secret_ref"] == "access_control.admin_password"
    assert access_control_payload["session_secret_secret_ref"] == "access_control.session_secret"
    assert "guest_password" not in access_control_payload
    assert "admin_password" not in access_control_payload
    assert "session_secret" not in access_control_payload
    env_content = env_path.read_text(encoding="utf-8")
    assert "TG_BOT_TOKEN=" not in env_content
    assert "TG_CHAT_ID=" not in env_content
    assert "TG_BOT_TOKEN_SECRET_REF=telegram.bot_token" in env_content
    assert "TG_CHAT_ID_SECRET_REF=telegram.chat_id" in env_content
    provider = EncryptedFileSecretProvider(settings.secrets_file, master_key=master_key)
    assert provider.get_secret("telegram.bot_token") == "bot-token"
    assert provider.get_secret("telegram.chat_id") == "chat-id"
    assert provider.get_secret("access_control.guest_password") == "guest-pass"
    assert provider.get_secret("access_control.admin_password") == "admin-pass-next"
    assert provider.get_secret("access_control.session_secret") == "dev-session-secret-1234567890"


def test_import_excel_endpoint_settings_only_keeps_existing_accounts(monkeypatch, tmp_path: Path) -> None:
    master_key = create_master_key()
    env_path = tmp_path / ".env"
    env_path.write_text("TG_ENABLED=true\n", encoding="utf-8")
    settings = Settings(
        _env_file=None,
        monitor_accounts_file=tmp_path / "config" / "binance_monitor_accounts.json",
        access_control_config_file=tmp_path / "access_control.json",
        secrets_file=tmp_path / "config" / "secrets.enc.json",
        env_file_path=env_path,
        monitor_master_key=master_key,
        monitor_refresh_interval_ms=999999,
        monitor_history_window_days=3,
        monitor_history_db_path=tmp_path / "history.db",
    )
    settings.monitor_accounts_file.parent.mkdir(parents=True, exist_ok=True)
    existing_payload = {
        "main_accounts": [
            {
                "main_id": "existing",
                "name": "Existing",
                "children": [
                    {
                        "account_id": "sub1",
                        "name": "Old",
                        "api_key_secret_ref": "accounts.existing.sub1.api_key",
                        "api_secret_secret_ref": "accounts.existing.sub1.api_secret",
                    }
                ],
            }
        ]
    }
    settings.monitor_accounts_file.write_text(json.dumps(existing_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    provider = EncryptedFileSecretProvider(settings.secrets_file, master_key=master_key)
    provider.set_secret("accounts.existing.sub1.api_key", "k1")
    provider.set_secret("accounts.existing.sub1.api_secret", "s1")
    monkeypatch.setattr(api_module, "settings", settings)

    with TestClient(api_module.app) as client:
        client.app.state.access_control = _build_disabled_access_control(tmp_path, settings)
        controller = AccountMonitorController(settings, gateway_factory=lambda account: FakeImportGateway(account))
        client.app.state.monitor = controller
        try:
            response = client.post(
                "/api/config/import/excel",
                files={
                    "file": (
                        "accounts.xlsx",
                        _build_workbook_bytes(
                            account_rows=[
                                ["main_id", "main_name", "account_id", "name", "api_key", "api_secret", "uid", "use_testnet"],
                            ],
                            settings_rows=[
                                ["key", "value", "notes"],
                                ["access_control.guest_password", "guest-pass-next", "rotate guest password"],
                            ],
                        ),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        finally:
            asyncio.run(controller.close())

    assert response.status_code == 200
    payload = response.json()
    assert payload["import_result"]["main_account_count"] == 0
    assert payload["import_result"]["account_count"] == 0
    assert payload["import_result"]["mode"] == "settings_only"
    assert payload["import_result"]["updated_settings_keys"] == ["access_control.guest_password"]
    assert payload["message"] == excel_import_settings_success_message()
    assert payload["refresh_result"] == {"success": True, "skipped": True, "message": ""}
    persisted = json.loads(settings.monitor_accounts_file.read_text(encoding="utf-8"))
    assert persisted == existing_payload
    provider = EncryptedFileSecretProvider(settings.secrets_file, master_key=master_key)
    assert provider.get_secret("accounts.existing.sub1.api_key") == "k1"
    assert provider.get_secret("accounts.existing.sub1.api_secret") == "s1"
    assert provider.get_secret("access_control.guest_password") == "guest-pass-next"


def test_import_excel_endpoint_settings_only_reloads_telegram_credentials(monkeypatch, tmp_path: Path) -> None:
    master_key = create_master_key()
    env_path = tmp_path / ".env"
    env_path.write_text("TG_ENABLED=true\n", encoding="utf-8")
    settings = Settings(
        _env_file=None,
        monitor_accounts_file=tmp_path / "config" / "binance_monitor_accounts.json",
        access_control_config_file=tmp_path / "access_control.json",
        secrets_file=tmp_path / "config" / "secrets.enc.json",
        env_file_path=env_path,
        monitor_master_key=master_key,
        monitor_refresh_interval_ms=999999,
        monitor_history_window_days=3,
        monitor_history_db_path=tmp_path / "history.db",
        tg_enabled=True,
    )
    monkeypatch.setattr(api_module, "settings", settings)

    with TestClient(api_module.app) as client:
        client.app.state.access_control = _build_disabled_access_control(tmp_path, settings)
        controller = AccountMonitorController(settings, gateway_factory=lambda account: FakeImportGateway(account))
        client.app.state.monitor = controller
        try:
            assert controller._telegram_notifications.enabled is False
            response = client.post(
                "/api/config/import/excel",
                files={
                    "file": (
                        "accounts.xlsx",
                        _build_workbook_bytes(
                            account_rows=[
                                ["main_id", "main_name", "account_id", "name", "api_key", "api_secret", "uid", "use_testnet"],
                            ],
                            settings_rows=[
                                ["key", "value", "notes"],
                                ["telegram.bot_token", "bot-token-next", "rotate bot token"],
                                ["telegram.chat_id", "chat-id-next", "rotate chat id"],
                            ],
                        ),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        finally:
            asyncio.run(controller.close())

    assert response.status_code == 200
    assert controller._telegram_notifications.enabled is True
    assert controller._telegram_notifications._bot_token == "bot-token-next"
    assert controller._telegram_notifications._chat_id == "chat-id-next"


def test_import_excel_endpoint_accounts_only_does_not_require_optional_files(monkeypatch, tmp_path: Path) -> None:
    master_key = create_master_key()
    settings = Settings(
        _env_file=None,
        monitor_accounts_file=tmp_path / "config" / "binance_monitor_accounts.json",
        access_control_config_file=tmp_path / "missing" / "access_control.json",
        secrets_file=tmp_path / "config" / "secrets.enc.json",
        env_file_path=tmp_path / "missing" / ".env",
        monitor_master_key=master_key,
        monitor_refresh_interval_ms=999999,
        monitor_history_window_days=3,
        monitor_history_db_path=tmp_path / "history.db",
    )
    monkeypatch.setattr(api_module, "settings", settings)

    with TestClient(api_module.app) as client:
        client.app.state.access_control = _build_disabled_access_control(tmp_path / "runtime", settings)
        controller = AccountMonitorController(settings, gateway_factory=lambda account: FakeImportGateway(account))
        client.app.state.monitor = controller
        try:
            existing_access_control_payload = (
                json.loads(settings.access_control_config_file.read_text(encoding="utf-8-sig"))
                if settings.access_control_config_file.exists()
                else None
            )
            response = client.post(
                "/api/config/import/excel",
                files={
                    "file": (
                        "accounts.xlsx",
                        _build_workbook_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        finally:
            asyncio.run(controller.close())

    assert response.status_code == 200
    current_access_control_payload = (
        json.loads(settings.access_control_config_file.read_text(encoding="utf-8-sig"))
        if settings.access_control_config_file.exists()
        else None
    )
    assert current_access_control_payload == existing_access_control_payload
    assert Path(settings.env_file_path).exists() is False


def test_verify_project_secret_consistency_checks_existing_sources_independently(monkeypatch, tmp_path: Path) -> None:
    master_key = create_master_key()
    settings = Settings(
        _env_file=None,
        monitor_accounts_file=tmp_path / "config" / "binance_monitor_accounts.json",
        access_control_config_file=tmp_path / "missing" / "access_control.json",
        secrets_file=tmp_path / "config" / "secrets.enc.json",
        env_file_path=tmp_path / "missing" / ".env",
        monitor_master_key=master_key,
        monitor_history_db_path=tmp_path / "history.db",
    )
    settings.monitor_accounts_file.parent.mkdir(parents=True, exist_ok=True)
    settings.monitor_accounts_file.write_text(
        json.dumps(
            {
                "main_accounts": [
                    {
                        "main_id": "group_a",
                        "name": "Group A",
                        "children": [
                            {
                                "account_id": "sub1",
                                "name": "Sub One",
                                "api_key_secret_ref": "accounts.group_a.sub1.api_key",
                                "api_secret_secret_ref": "accounts.group_a.sub1.api_secret",
                            }
                        ],
                    }
                ]
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(api_module, "settings", settings)
    EncryptedFileSecretProvider(settings.secrets_file, master_key=master_key)

    with pytest.raises(RuntimeError, match="Secret store is missing refs"):
        api_module._verify_project_secret_consistency_if_ready()


def test_import_excel_endpoint_rotating_access_control_clears_session_cookie(monkeypatch, tmp_path: Path) -> None:
    master_key = create_master_key()
    env_path = tmp_path / ".env"
    env_path.write_text("TG_ENABLED=true\n", encoding="utf-8")
    settings = Settings(
        _env_file=None,
        monitor_accounts_file=tmp_path / "config" / "binance_monitor_accounts.json",
        access_control_config_file=tmp_path / "config" / "access_control.json",
        secrets_file=tmp_path / "config" / "secrets.enc.json",
        env_file_path=env_path,
        monitor_master_key=master_key,
        monitor_refresh_interval_ms=999999,
        monitor_history_window_days=3,
        monitor_history_db_path=tmp_path / "history.db",
    )
    monkeypatch.setattr(api_module, "settings", settings)

    with TestClient(api_module.app) as client:
        client.app.state.access_control = _build_enabled_access_control(tmp_path, settings)
        client.app.state.test_client_ip = "198.51.100.10"
        controller = AccountMonitorController(settings, gateway_factory=lambda account: FakeImportGateway(account))
        client.app.state.monitor = controller
        try:
            login = client.post("/api/auth/login", json={"password": "admin-pass"})
            assert login.status_code == 200
            csrf_token = login.json()["csrf_token"]
            assert client.cookies.get(AUTH_COOKIE_NAME)

            response = client.post(
                "/api/config/import/excel",
                headers={"X-CSRF-Token": csrf_token},
                files={
                    "file": (
                        "accounts.xlsx",
                        _build_workbook_bytes(
                            account_rows=[
                                ["main_id", "main_name", "account_id", "name", "api_key", "api_secret", "uid", "use_testnet"],
                            ],
                            settings_rows=[
                                ["key", "value", "notes"],
                                ["access_control.session_secret", "next-session-secret-value", "rotate session secret"],
                            ],
                        ),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            cookie_header = response.headers.get("set-cookie", "").lower()
            client_cookie = client.cookies.get(AUTH_COOKIE_NAME)
        finally:
            asyncio.run(controller.close())

    assert response.status_code == 200
    assert client_cookie is None
    assert AUTH_COOKIE_NAME in cookie_header
    assert "expires=" in cookie_header or "max-age=0" in cookie_header


def test_import_excel_endpoint_rolls_back_runtime_state_after_reload_failure(monkeypatch, tmp_path: Path) -> None:
    master_key = create_master_key()
    env_path = tmp_path / ".env"
    env_path.write_text("TG_ENABLED=true\n", encoding="utf-8")
    settings = Settings(
        _env_file=None,
        monitor_accounts_file=tmp_path / "config" / "binance_monitor_accounts.json",
        access_control_config_file=tmp_path / "access_control.json",
        secrets_file=tmp_path / "config" / "secrets.enc.json",
        env_file_path=env_path,
        monitor_master_key=master_key,
        monitor_refresh_interval_ms=999999,
        monitor_history_window_days=3,
        monitor_history_db_path=tmp_path / "history.db",
    )
    settings.monitor_accounts_file.parent.mkdir(parents=True, exist_ok=True)
    existing_payload = {
        "main_accounts": [
            {
                "main_id": "existing",
                "name": "Existing",
                "children": [
                    {
                        "account_id": "sub1",
                        "name": "Old",
                        "api_key_secret_ref": "accounts.existing.sub1.api_key",
                        "api_secret_secret_ref": "accounts.existing.sub1.api_secret",
                    }
                ],
            }
        ]
    }
    settings.monitor_accounts_file.write_text(json.dumps(existing_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    provider = EncryptedFileSecretProvider(settings.secrets_file, master_key=master_key)
    provider.set_secret("accounts.existing.sub1.api_key", "k1")
    provider.set_secret("accounts.existing.sub1.api_secret", "s1")
    monkeypatch.setattr(api_module, "settings", settings)

    with TestClient(api_module.app, raise_server_exceptions=False) as client:
        access_control = _build_disabled_access_control(tmp_path, settings)
        client.app.state.access_control = access_control
        controller = ReloadThenFailController(settings, gateway_factory=lambda account: FakeImportGateway(account))
        client.app.state.monitor = controller
        try:
            response = client.post(
                "/api/config/import/excel",
                files={
                    "file": (
                        "accounts.xlsx",
                        _build_workbook_bytes(
                            settings_rows=[
                                ["key", "value", "notes"],
                                ["access_control.guest_password", "guest-pass-next", "rotate guest password"],
                            ],
                        ),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        finally:
            asyncio.run(controller.close())

    assert response.status_code == 500
    assert response.json()["detail"] == "reload failed after state change"
    persisted = json.loads(settings.monitor_accounts_file.read_text(encoding="utf-8"))
    assert persisted == existing_payload
    assert sorted(settings.monitor_accounts) == ["existing.sub1"]
    assert access_control.config is not None
    assert access_control.config.guest_password == "guest-pass"
    provider = EncryptedFileSecretProvider(settings.secrets_file, master_key=master_key)
    assert provider.get_secret("accounts.existing.sub1.api_key") == "k1"
    assert provider.get_secret("accounts.existing.sub1.api_secret") == "s1"


def test_import_excel_endpoint_rolls_back_runtime_telegram_settings_after_reload_failure(
    monkeypatch, tmp_path: Path
) -> None:
    master_key = create_master_key()
    env_path = tmp_path / ".env"
    env_path.write_text("TG_ENABLED=true\n", encoding="utf-8")
    settings = Settings(
        _env_file=None,
        monitor_accounts_file=tmp_path / "config" / "binance_monitor_accounts.json",
        access_control_config_file=tmp_path / "access_control.json",
        secrets_file=tmp_path / "config" / "secrets.enc.json",
        env_file_path=env_path,
        monitor_master_key=master_key,
        monitor_refresh_interval_ms=999999,
        monitor_history_window_days=3,
        monitor_history_db_path=tmp_path / "history.db",
        tg_enabled=True,
        tg_bot_token_secret_ref="telegram.bot_token.process",
        tg_chat_id_secret_ref="telegram.chat_id.process",
    )
    provider = EncryptedFileSecretProvider(settings.secrets_file, master_key=master_key)
    provider.set_secret("telegram.bot_token.process", "process-bot-token")
    provider.set_secret("telegram.chat_id.process", "process-chat-id")
    monkeypatch.setattr(api_module, "settings", settings)

    with TestClient(api_module.app, raise_server_exceptions=False) as client:
        client.app.state.access_control = _build_disabled_access_control(tmp_path, settings)
        controller = ReloadThenFailController(settings, gateway_factory=lambda account: FakeImportGateway(account))
        client.app.state.monitor = controller
        try:
            assert controller._telegram_notifications.enabled is True
            assert controller._telegram_notifications._bot_token == "process-bot-token"
            assert controller._telegram_notifications._chat_id == "process-chat-id"

            response = client.post(
                "/api/config/import/excel",
                files={
                    "file": (
                        "accounts.xlsx",
                        _build_workbook_bytes(
                            settings_rows=[
                                ["key", "value", "notes"],
                                ["telegram.bot_token", "bot-token-next", "rotate bot token"],
                                ["telegram.chat_id", "chat-id-next", "rotate chat id"],
                            ],
                        ),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        finally:
            asyncio.run(controller.close())

    assert response.status_code == 500
    assert response.json()["detail"] == "reload failed after state change"
    assert controller._telegram_notifications.enabled is True
    assert controller._telegram_notifications._bot_token == "process-bot-token"
    assert controller._telegram_notifications._chat_id == "process-chat-id"
    assert "TG_BOT_TOKEN_SECRET_REF" not in env_path.read_text(encoding="utf-8")
    assert "TG_CHAT_ID_SECRET_REF" not in env_path.read_text(encoding="utf-8")


def test_import_excel_endpoint_rejects_non_xlsx(monkeypatch, tmp_path: Path) -> None:
    settings = Settings(
        _env_file=None,
        monitor_accounts_file=tmp_path / "config" / "binance_monitor_accounts.json",
        access_control_config_file=tmp_path / "access_control.json",
        env_file_path=tmp_path / ".env",
        monitor_history_db_path=tmp_path / "history.db",
    )
    monkeypatch.setattr(api_module, "settings", settings)

    with TestClient(api_module.app) as client:
        client.app.state.access_control = _build_disabled_access_control(tmp_path, settings)
        response = client.post(
            "/api/config/import/excel",
            files={"file": ("accounts.csv", b"not-used", "text/csv")},
        )

    assert response.status_code == 400
    assert response.json()["detail"] == "Only .xlsx files are supported"


def test_import_excel_validation_failure_does_not_overwrite_existing_file(monkeypatch, tmp_path: Path) -> None:
    master_key = create_master_key()
    env_path = tmp_path / ".env"
    env_path.write_text("TG_ENABLED=true\n", encoding="utf-8")
    settings = Settings(
        _env_file=None,
        monitor_accounts_file=tmp_path / "config" / "binance_monitor_accounts.json",
        access_control_config_file=tmp_path / "access_control.json",
        secrets_file=tmp_path / "config" / "secrets.enc.json",
        env_file_path=env_path,
        monitor_master_key=master_key,
        monitor_history_db_path=tmp_path / "history.db",
    )
    settings.monitor_accounts_file.parent.mkdir(parents=True, exist_ok=True)
    existing_payload = {
        "main_accounts": [
            {
                "main_id": "existing",
                "name": "Existing",
                "children": [
                    {
                        "account_id": "sub1",
                        "name": "Old",
                        "api_key_secret_ref": "accounts.existing.sub1.api_key",
                        "api_secret_secret_ref": "accounts.existing.sub1.api_secret",
                    }
                ],
            }
        ]
    }
    settings.monitor_accounts_file.write_text(json.dumps(existing_payload), encoding="utf-8")
    provider = EncryptedFileSecretProvider(settings.secrets_file, master_key=master_key)
    provider.set_secret("accounts.existing.sub1.api_key", "k1")
    provider.set_secret("accounts.existing.sub1.api_secret", "s1")
    monkeypatch.setattr(api_module, "settings", settings)

    with TestClient(api_module.app) as client:
        client.app.state.access_control = _build_disabled_access_control(tmp_path, settings)
        response = client.post(
            "/api/config/import/excel",
            files={"file": ("accounts.xlsx", _build_invalid_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    assert response.status_code == 400
    assert response.json()["detail"] == "Missing required columns: api_secret"
    persisted = json.loads(settings.monitor_accounts_file.read_text(encoding="utf-8"))
    assert persisted == existing_payload
    provider = EncryptedFileSecretProvider(settings.secrets_file, master_key=master_key)
    assert provider.get_secret("accounts.existing.sub1.api_key") == "k1"


def test_import_excel_endpoint_reports_refresh_failure_after_successful_write(monkeypatch, tmp_path: Path) -> None:
    master_key = create_master_key()
    env_path = tmp_path / ".env"
    env_path.write_text("TG_ENABLED=true\n", encoding="utf-8")
    settings = Settings(
        _env_file=None,
        monitor_accounts_file=tmp_path / "config" / "binance_monitor_accounts.json",
        access_control_config_file=tmp_path / "access_control.json",
        secrets_file=tmp_path / "config" / "secrets.enc.json",
        env_file_path=env_path,
        monitor_master_key=master_key,
        monitor_refresh_interval_ms=999999,
        monitor_history_window_days=3,
        monitor_history_db_path=tmp_path / "history.db",
    )
    monkeypatch.setattr(api_module, "settings", settings)

    with TestClient(api_module.app) as client:
        client.app.state.access_control = _build_disabled_access_control(tmp_path, settings)
        controller = AccountMonitorController(settings, gateway_factory=lambda account: FailingImportGateway(account))
        client.app.state.monitor = controller
        try:
            response = client.post(
                "/api/config/import/excel",
                files={"file": ("accounts.xlsx", _build_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        finally:
            asyncio.run(controller.close())

    assert response.status_code == 200
    payload = response.json()
    assert payload["message"] == excel_import_refresh_failed_message()
    assert payload["import_result"]["main_account_count"] == 2
    assert payload["refresh_result"]["success"] is False
    assert payload["service"]["account_ids"] == ["group_a.sub1", "group_a.sub2", "group_b.sub1"]
    persisted = json.loads(settings.monitor_accounts_file.read_text(encoding="utf-8"))
    assert len(persisted["main_accounts"]) == 2
    assert persisted["main_accounts"][0]["children"][0]["api_key_secret_ref"] == "accounts.group_a.sub1.api_key"


def test_download_excel_template_endpoint_returns_header_only_workbook(tmp_path: Path) -> None:
    temp_settings = Settings(
        _env_file=None,
        access_control_config_file=tmp_path / "access_control.json",
        env_file_path=tmp_path / ".env",
    )
    with TestClient(api_module.app) as client:
        client.app.state.access_control = _build_disabled_access_control(tmp_path, temp_settings)
        response = client.get("/api/config/import/excel-template")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response.headers["content-disposition"] == 'attachment; filename="monitor_accounts_template.xlsx"'

    loaded = _load_workbook_bytes(response.content)
    try:
        worksheet = loaded["accounts"]
        settings_sheet = loaded["settings"]
        guide_rows = list(loaded["guide"].iter_rows(values_only=True))
        account_rows = list(worksheet.iter_rows(values_only=True))
        settings_rows = list(settings_sheet.iter_rows(values_only=True))
    finally:
        loaded.close()

    assert account_rows == [(
        "main_id",
        "main_name",
        "account_id",
        "name",
        "api_key",
        "api_secret",
        "uid",
        "use_testnet",
        "rest_base_url",
        "ws_base_url",
    )]
    assert settings_rows[0] == ("key", "value", "notes")
    assert ("template_version", TEMPLATE_VERSION) in guide_rows
    assert any(row[0] == "cleanup" and "delete the local Excel file" in row[1] for row in guide_rows if row[0])
    assert any(row[0] == "telegram.bot_token" for row in settings_rows[1:])


def _build_workbook_bytes(
    *,
    account_rows: list[list[object]] | None = None,
    settings_rows: list[list[object]] | None = None,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "accounts"
    for row in account_rows or [
        ["main_id", "main_name", "account_id", "name", "api_key", "api_secret", "uid", "use_testnet"],
        ["group_a", "Group A", "main", "Main Transfer", "mk1", "ms1", "123456789", ""],
        ["group_a", "Group A", "sub1", "Sub One", "k1", "s1", "223456789", "true"],
        ["group_a", "Group A", "sub2", "Sub Two", "k2", "s2", "", ""],
        ["group_b", "Group B", "sub1", "Sub Three", "k3", "s3", "323456789", "false"],
    ]:
        worksheet.append(row)

    settings_sheet = workbook.create_sheet("settings")
    settings_sheet.append(["key", "value", "notes"])
    if settings_rows:
        for row in settings_rows[1:] if settings_rows and settings_rows[0][:2] == ["key", "value"] else settings_rows:
            settings_sheet.append(row)

    guide = workbook.create_sheet("guide")
    guide.append(["section", "content"])
    guide.append(["template_version", TEMPLATE_VERSION])

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _build_invalid_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "accounts"
    worksheet.append(["main_id", "main_name", "account_id", "name", "api_key"])
    worksheet.append(["group_a", "Group A", "sub1", "Sub One", "k1"])
    guide = workbook.create_sheet("guide")
    guide.append(["section", "content"])
    guide.append(["template_version", TEMPLATE_VERSION])
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _load_workbook_bytes(content: bytes):
    from openpyxl import load_workbook

    return load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
